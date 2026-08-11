"""Nhập kịch bản từ CSV/Excel — 6 cột: Mã block, Thời lượng, Loại Visual,
Hình ảnh & Hiệu ứng (Visual/FX), Âm thanh & Nhạc nền (Audio/SFX), Kịch bản Giọng đọc (VO).

Quyết định triển khai (đã build vòng 4, khác design gốc): design parse file bằng
SheetJS (`xlsx.full.min.js`, ~800KB) chạy phía client. Bản build này parse phía
SERVER (Python) — tránh nạp thêm thư viện JS nặng vào bundle Electron, gom logic
parse/validate vào 1 chỗ và test được bằng pytest (đúng nguyên tắc đã áp dụng cho
toàn bộ backend). Frontend chỉ upload file thô, không tự parse.
"""
from __future__ import annotations

import csv
import io
import re

COLUMN_KEYWORDS: dict[str, list[str]] = {
    "block_id": ["mã"],
    "ts": ["thời lượng"],
    "visual_type": ["loại visual"],
    "visual": ["hình ảnh", "visual/fx"],
    "audio_sfx": ["âm thanh", "audio/sfx"],
    "vo": ["kịch bản", "vo content", "giọng đọc"],
}

REQUIRED_COLUMNS_MESSAGE = (
    "File cần đủ 6 cột: Mã block, Thời lượng, Loại Visual, "
    "Hình ảnh & Hiệu ứng (Visual/FX), Âm thanh & Nhạc nền (Audio/SFX), "
    "Kịch bản Giọng đọc (VO Content)."
)

DEFAULT_BLOCK_SEC = 8  # ước tính khi không đọc được "Thời lượng" của 1 block


class ScriptImportError(Exception):
    """Lỗi hiển thị trực tiếp cho người dùng trong dialog xác nhận nhập kịch bản."""


def _find_col(header: list[str], keywords: list[str]) -> int:
    for i, h in enumerate(header):
        low = h.lower().strip()
        if any(k in low for k in keywords):
            return i
    return -1


def _rows_from_csv(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader if any(str(c).strip() for c in row)]


def _rows_from_xlsx(content: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        cells = ["" if c is None else str(c) for c in row]
        if any(c.strip() for c in cells):
            rows.append(cells)
    return rows


_TS_RANGE = re.compile(r"^(\d{1,2}):(\d{2})\s*[–-]\s*(\d{1,2}):(\d{2})$")
_TS_SINGLE = re.compile(r"^(\d{1,2}):(\d{2})$")


def parse_ts(ts: str) -> tuple[int | None, int | None]:
    """'0:00–0:05' -> (0, 5). '1:30' -> (90, None). Không nhận dạng được -> (None, None)."""
    ts = ts.strip()
    m = _TS_RANGE.match(ts)
    if m:
        start = int(m.group(1)) * 60 + int(m.group(2))
        end = int(m.group(3)) * 60 + int(m.group(4))
        return start, end
    m = _TS_SINGLE.match(ts)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2)), None
    return None, None


def parse_script_rows(rows: list[list[str]]) -> dict:
    """Trả `{beats, stats, full_text}`. Raise `ScriptImportError` (thông điệp tiếng
    Việt hiển thị thẳng cho người dùng) nếu file trống/thiếu cột/không có block."""
    if not rows:
        raise ScriptImportError("File trống — không tìm thấy dữ liệu.")

    header = [str(h or "") for h in rows[0]]
    idx = {k: _find_col(header, kws) for k, kws in COLUMN_KEYWORDS.items()}
    missing = [k for k, v in idx.items() if v == -1]
    if missing:
        raise ScriptImportError(REQUIRED_COLUMNS_MESSAGE)

    data_rows = rows[1:]
    if not data_rows:
        raise ScriptImportError("File không có block nào ở dưới dòng tiêu đề.")

    def cell(r: list, i: int) -> str:
        return str(r[i]).strip() if i < len(r) and r[i] is not None else ""

    beats = []
    cursor = 0
    for r in data_rows:
        ts_raw = cell(r, idx["ts"])
        start, end = parse_ts(ts_raw)
        if start is None:
            start = cursor
        beats.append(
            {
                "block_id": cell(r, idx["block_id"]),
                "ts_label": ts_raw or f"{start // 60}:{start % 60:02d}",
                "timestamp_sec": start,
                "end_sec": end,
                "visual_type": cell(r, idx["visual_type"]),
                "visual": cell(r, idx["visual"]),
                "direction": cell(r, idx["audio_sfx"]),
                "direction_label": "Audio/SFX",
                "audio": cell(r, idx["vo"]),
                "anchor": False,
            }
        )
        cursor = (end if end is not None else start) + DEFAULT_BLOCK_SEC

    full_text = "\n\n".join(b["audio"] for b in beats if b["audio"])
    word_count = len(full_text.split())
    has_duration = any(b["end_sec"] is not None for b in beats)
    if has_duration:
        total_sec = sum((b["end_sec"] - b["timestamp_sec"]) for b in beats if b["end_sec"] is not None)
        duration_label = f"{total_sec // 60}:{total_sec % 60:02d}"
    else:
        duration_label = "Không xác định"

    return {
        "beats": beats,
        "stats": {"block_count": len(beats), "word_count": word_count, "duration_label": duration_label},
        "full_text": full_text,
    }


def parse_script_file(content: bytes, filename: str) -> dict:
    is_csv = filename.lower().endswith(".csv")
    try:
        rows = _rows_from_csv(content) if is_csv else _rows_from_xlsx(content)
    except ScriptImportError:
        raise
    except Exception as e:  # noqa: BLE001
        raise ScriptImportError("Không đọc được file. Kiểm tra định dạng CSV/Excel và thử lại.") from e
    return parse_script_rows(rows)
